#-*- coding:utf8 -*-
import requests
import os
from openpyxl import load_workbook
from openpyxl import Workbook

def get_json(url, page, position_name):
	data = {'first': 'true', 'pn': page, 'kd': position_name}
	headers = {
            'Accept':'application/json, text/javascript, */*; q=0.01',
            'Accept-Encoding':'gzip, deflate, br',
            'Accept-Language':'zh-CN,zh;q=0.8',
            'Cache-Control':'max-age=0',
            'Connection':'keep-alive',
            'Content-Length':'0',
            'Cookie':'user_trace_token=20160919110304-aaa88f9d66cf47c7a2828f45af41f4da; LGUID=20160919110304-952bc0a3-7e15-11e6-bb00-525400f775ce; gr_user_id=dcc7b4ed-b22d-4c0b-91e6-877a38a4a956; yun_sso_auth=21d6d98169560b66f83023c749e0c31e3176407836ba9ea810fbc78709f1f1fffc5ebf1ea50c37e20d1403724c92c533d8d32d9459ebf9d4d591f192a782c3c6a5f10650f04cb1bfc6d769721adcdb380b7801fce3edd89a; yun_sso_logout=false; TG-TRACK-CODE=gongsi_logo; JSESSIONID=CE38054580F509E8E72BE9B947F1869F; _gat=1; PRE_UTM=; PRE_HOST=; PRE_SITE=https%3A%2F%2Fwww.lagou.com%2Fgongsi%2F; PRE_LAND=https%3A%2F%2Fwww.lagou.com%2Fgongsi%2F2-0-0; _putrc=91F120CB74730D20; login=true; unick=hh; showExpriedIndex=1; showExpriedCompanyHome=1; showExpriedMyPublish=1; hasDeliver=0; index_location_city=%E5%8C%97%E4%BA%AC; Hm_lvt_4233e74dff0ae5bd0a3d81c6ccf756e6=1487217558,1487733880,1487831545,1487904174; Hm_lpvt_4233e74dff0ae5bd0a3d81c6ccf756e6=1487928012; _ga=GA1.2.964704155.1474254184; LGSID=20170224172009-705bb23c-fa72-11e6-9018-5254005c3644; LGRID=20170224172011-716046c3-fa72-11e6-9018-5254005c3644',
            'Host':'www.lagou.com',
            'Origin':'https://www.lagou.com',
            'Referer':'https://www.lagou.com/gongsi/2-0-0',
            'User-Agent':'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/56.0.2924.87 Safari/537.36',
            'X-Anit-Forge-Code':'0',
            'X-Anit-Forge-Token':'None',
            'X-Requested-With':'XMLHttpRequest'
    }
	json = requests.post(url, data, headers = headers).json()
	list_con = json['content']['positionResult']['result']
	info_list = []
	for i in list_con:
		info = []
		info.append(i['companyShortName'])
		info.append(i['companyFullName'])
		info.append(i['positionName'])
		info.append(i['salary'])
		info.append(i['city'])
		info.append(i['district'])
		info.append(i['workYear'])
		info.append(i['education'])
		info_list.append(info)
	return info_list

def main():
	position_name = input(u'职位名：')
	page = 1
	url = 'https://www.lagou.com/jobs/positionAjax.json?px=default&city=%E5%8C%97%E4%BA%AC&needAddtionalResult=false'
	info_result = []
	while page < 31:
		print('----正在爬取第%d页----'%page)
		info = get_json(url, page, position_name)
		info_result = info_result + info
		page += 1
	if not os.path.exists('职位信息.xlsx'):
		wb = Workbook()
		ws = wb.active
		ws.title = position_name
	else:
		wb = load_workbook(filename = '职位信息.xlsx')
		ws = wb.create_sheet(title = position_name)
	for row in info_result:
		ws.append(row)
	wb.save('职位信息.xlsx')
	print('爬取完毕')

if __name__ == '__main__':
	main()
