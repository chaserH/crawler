#!/usr/bin/python
# -*- coding: utf-8 -*-

import os
#import json
import requests
import datetime
from pyquery import PyQuery as pq
from openpyxl import Workbook
from openpyxl import load_workbook


#获取城市ID列表
def get_cityId_list(url):
    cityId_list = []
    html = pq(url= url)
    for areaId in html.find('#filterCollapse').find('div[class="has-more workcity"]').eq(0).find('div[class="more more-positions"]').find("a[data-lg-tj-cid='idnull']"):
        aId = pq(areaId).attr('data-id')
        if(aId=='0'):
            continue
        cityId_list.append(aId)
    return cityId_list

#获取城市名称列表
def get_city_name_list(url):
    city_name_list = []
    url = 'https://www.lagou.com/gongsi/'
    html = pq(url=url)
    for areaId in html.find('#filterCollapse').find('div[class="has-more workcity"]').eq(0).find('div[class="more more-positions"]').find("a[data-lg-tj-cid='idnull']"):
        area_name=pq(areaId).html()
        if area_name=="全国":
            continue
        city_name_list.append(area_name)
    return city_name_list

#获取城市下一共有多少页
def get_city_page(aId,page_num):
    try:
        param = {'first': 'false', 'pn': page_num, 'sortField': '0', 'havemark': '0'} #访问参数
        r = urllib.request.urlopen('https://www.lagou.com/gongsi/'+aId+'-0-0') #requsets请求
        page_num += 1
        if(len(r.json()['result'])/16==1):
            return get_city_page(aId,page_num)
        else:
            return page_num
    except:
        return page_num-1

#根据城市ID获取所有公司信息
def get_company_list(areaId):
    company_list = []
    city_page_total=get_city_page(areaId,1)
    for pageIndex in range(1,city_page_total+1):
        print('正在爬取第'+str(pageIndex)+'页')
        json_url = 'https://www.lagou.com/gongsi/'+areaId+'-0-0'
        param = {'first': 'false', 'pn': str(pageIndex), 'sortField': '0', 'havemark': '0'} #访问参数
        headers = {
            'Accept':'application/json, text/javascript, */*; q=0.01',
            'Accept-Encoding':'gzip, deflate, br',
            'Accept-Language':'zh-CN,zh;q=0.8',
            'Cache-Control':'max-age=0',
            'Connection':'keep-alive',
            'Content-Length':'0',
            'Cookie':'user_trace_token=20160919110304-aaa88f9d66cf47c7a2828f45af41f4da; LGUID=20160919110304-952bc0a3-7e15-11e6-bb00-525400f775ce; gr_user_id=dcc7b4ed-b22d-4c0b-91e6-877a38a4a956; yun_sso_auth=21d6d98169560b66f83023c749e0c31e3176407836ba9ea810fbc78709f1f1fffc5ebf1ea50c37e20d1403724c92c533d8d32d9459ebf9d4d591f192a782c3c6a5f10650f04cb1bfc6d769721adcdb380b7801fce3edd89a; yun_sso_logout=false; TG-TRACK-CODE=gongsi_logo; JSESSIONID=CE38054580F509E8E72BE9B947F1869F; _gat=1; PRE_UTM=; PRE_HOST=; PRE_SITE=https%3A%2F%2Fwww.lagou.com%2Fgongsi%2F; PRE_LAND=https%3A%2F%2Fwww.lagou.com%2Fgongsi%2F2-0-0; _putrc=91F120CB74730D20; login=true; unick=hh; showExpriedIndex=1; showExpriedCompanyHome=1; showExpriedMyPublish=1; hasDeliver=0; index_location_city=%E5%8C%97%E4%BA%AC; Hm_lvt_4233e74dff0ae5bd0a3d81c6ccf756e6=1487217558,1487733880,1487831545,1487904174; Hm_lpvt_4233e74dff0ae5bd0a3d81c6ccf756e6=1487928012; _ga=GA1.2.964704155.1474254184; LGSID=20170224172009-705bb23c-fa72-11e6-9018-5254005c3644; LGRID=20170224172011-716046c3-fa72-11e6-9018-5254005c3644',
            'Host':'www.lagou.com',
            'Origin':'https://www.lagou.com',
            'Referer':'https://www.lagou.com/gongsi/2-0-0',
            'User-Agent':'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/56.0.2924.87 Safari/537.36',
            'X-Anit-Forge-Code':'0',
            'X-Anit-Forge-Token':'None',
            'X-Requested-With':'XMLHttpRequest'
        }
        r = requests.post(json_url, params=param, headers = headers) #requsets请求
        #text = r.read()
        #text = text.replace("'","\"")
        msg = r.json()
        #msg = json.loads(r.text)
        try:
            for company in msg['result']:
               company_list.append([company['city'],company['cityScore'],company['companyFeatures'],company['companyId'],company['companyLabels'],company['companyLogo'],company['companyName'],str(company['companyPositions']),company['companyShortName'],company['countryScore'],company['createTime'],company['finaceStage'],company['industryField'],company['interviewRemarkNum'],company['otherLabels'], company['positionNum'],company['processRate'],str(datetime.datetime.now())])
        except:
            print('爬取编号为'+str(areaId)+'城市时第'+str(pageIndex)+'页出现了错误,错误时请求返回内容为：'+str(msg))
            continue
    return company_list

#写入Excel文件方法
def write_file(fileName):
    list = []
    wb = Workbook()
    #ws = wb.active
    url = 'https://www.lagou.com/gongsi/'
    area_name_list = get_city_name_list(url)
    for area_name in area_name_list:
        wb.create_sheet(title = area_name)
        file_name = fileName+'.xlsx'
        wb.save(file_name)
    areaId_list = get_cityId_list(url)
    for areaId in areaId_list:
        company_list = get_company_list(areaId)
        print('正在爬取----->****'+company_list[0][0]+'****公司列表')
        wb1 = load_workbook(file_name)
        ws = wb1.get_sheet_by_name(company_list[0][0])
        ws.append(['城市名称','城市得分','公司期望','公司ID','公司标签','公司Logo','发展阶段','企业名称','企业位置','企业简称','注册时间','财务状况','行业','在招职位','其他标签','简历处理率'])
    for company in company_list:
        ws.append([company[0],str(company[1]),company[2],str(company[3]),company[4],company[5],company[6],company[7],company[8],company[9],company[10],company[11],company[12],company[13],company[14],company[15]])
        wb1.save(file_name)


#print('请输入文件名称'.encode('utf-8'))
#file_name = input(u'please input filename: '.encode('utf-8').decode('unicode_escape'))
file_name = 'lagou'
print(str(datetime.datetime.now()))
write_file(file_name)
print(str(datetime.datetime.now()))

